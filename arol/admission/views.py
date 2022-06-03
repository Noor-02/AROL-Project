from rest_framework import viewsets
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.pagination import PageNumberPagination
from django.utils.timezone import now
from .models import (
    Application,
    Education_Detail,
    Employment,
    Profile,
    Project_Detail,
    Qualifying_Examination,
    Recommendation,
)
from .serializers import (
    Application_Serializer,
    Education_Serializer,
    Employment_Serializer,
    Examination_Serializer,
    Profile_Serializer,
    Project_Serializer,
    Recommendation_Serializer,
)


class Application_Viewset(viewsets.ModelViewSet):
    """
    Returns a list of all **active** accounts in the system.
    For more details on how accounts are activated please [see here][ref].

    [ref]: http://example.com/activating-accounts
    """

    serializer_class = Application_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Application.objects.all()


class Education_Viewset(viewsets.ModelViewSet):
    serializer_class = Education_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Education_Detail.objects.all()


class Employment_Viewset(viewsets.ModelViewSet):
    serializer_class = Employment_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Employment.objects.all()


class Examination_Viewset(viewsets.ModelViewSet):
    serializer_class = Examination_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Qualifying_Examination.objects.all()


class Profile_Viewset(viewsets.ModelViewSet):
    serializer_class = Profile_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Profile.objects.all()


class Project_Viewset(viewsets.ModelViewSet):
    serializer_class = Project_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Project_Detail.objects.all()


class Recommendation_Viewset(viewsets.ModelViewSet):
    serializer_class = Recommendation_Serializer
    pagination_class = PageNumberPagination
    filter_backends = (SearchFilter, OrderingFilter)
    # search_fields = ["s_no", "name", "occupation"]

    def get_queryset(self):
        return Recommendation.objects.all()


from openpyxl import Workbook
from django.http import HttpResponse


def export_xlsx(request):
    queryset = Profile.objects.all()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response[
        "Content-Disposition"
    ] = "attachment; filename={year}-Profiles.xlsx".format(
        year=now().strftime("%y"),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Profiles"

    columns = [
        "Applicant ID",
        "Account",
        "Indian Applicant",
        "Nationality",
        "Full Name",
        "Father's/Spouse Name",
        "Marital Status",
        "Date of Birth",
        "Gender",
        "Caste Category",
        "Contact Number",
        "Parent Contact Number",
        "PwD",
        "Type of Disability",
        "Address",
        "City",
        "State",
        "Pin/Zip",
        "Address",
        "City",
        "State",
        "Pin/Zip",
    ]
    row_num = 1

    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for profile in queryset:
        row_num += 1

        # Define the data for each cell in the row
        print(type(profile.account))
        row = [
            profile.applicant_id,
            profile.account,
            profile.type_of_applicant,
            profile.nationality,
            profile.full_name,
            profile.father_or_spouse_name,
            profile.marital_status,
            profile.date_of_birth,
            profile.gender,
            profile.caste_category,
            profile.contact_number,
            profile.parent_contact_number,
            profile.pwd,
            profile.disability,
            profile.c_address,
            profile.c_city,
            profile.c_state,
            profile.c_pin,
            profile.p_address,
            profile.p_city,
            profile.p_state,
            profile.p_pin,
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)

    workbook.save(response)

    return response
