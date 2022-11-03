from django.http import HttpResponse
from openpyxl import Workbook

from ..models import Application,Profile,Education_Detail,Employment,Project_Detail,Recommendation
from ..permissions import IsStaff

def generate_xlsx(request, querystring,form):

    IsStaff(request)
    personalProfileRequired=False
    educationalDetailsRequired=False
    employmentDetailsRequired=False
    projectDetailsRequired=False     
    #print(form.cleaned_data['my_field'])


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=applications.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Applications"
    columns = [
        "Application ID",
        "Applicant ID",
        "Advertisement ID",
        "Payment ID",
        "Approved",
        "Date of Application",
    ]

    li = list(querystring.split("'"))
    n=len(li)
    queryset=[]

    for i in range(1,n-1,2):
        queryset.append(li[i])

    row_num = 1
    applicantID=[]

    for col_num, column_title in enumerate(columns, start=1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    row_num = 3

    for query in queryset:
        application=Application.objects.get(application_id=query)
        row =[
            application.application_id,
            application.applicant_id,
            application.advertisement_id,
            application.payment_id,
            application.is_approved,
            application.date_applied,
        ]
        applicantID.append(application.applicant_id)
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)
        row_num=row_num+1

    col=len(columns)+1

    if 'personal_profile' in form:
        required_columns=form.getlist('personal_profile')
        personalProfileColumns=[
            "Full Name",
            "Father's/Spouse Name",
            "Date of Birth",
            "Indian Applicant",
            "Nationality",
            "Maritial Status",
            "Gender",
            "Category",
            "Contact Number",
            "Parent Contact Number",
            "Correspondence Address",
            "Correspondence City",
            "Correspondence State",
            "Correspondence Pin/Zip",
            "Permanent Address",
            "Permanent City",
            "Permanent State",
            "Permanent Pin/Zip",
            "Person With Disability"
        ]
        if int(required_columns[0])==0:
            required_columns=[]
            for i in range(1,len(personalProfileColumns)+1):
                required_columns.append(i)
        row_num = 1
        for i in range(0,len(required_columns)):
            cell=worksheet.cell(row=row_num,column=col+i)
            cell.value=personalProfileColumns[int(required_columns[i])-1]
        row_num=3
        for query in applicantID:
            profile=Profile.objects.get(applicant_id=query)
            row =[
                profile.full_name,
                profile.father_or_spouse_name,
                profile.date_of_birth,
                profile.type_of_applicant,
                profile.nationality,
                profile.marital_status,
                profile.gender,
                profile.category,
                profile.contact_number,
                profile.parent_contact_number,
                profile.c_address,
                profile.c_city,
                profile.c_state,
                profile.c_pin,
                profile.p_address,
                profile.p_city,
                profile.p_state,
                profile.p_pin,
                profile.pwd,
            ]
            for i in range(0,len(required_columns)):
                cell=worksheet.cell(row=row_num,column=col+i)
                cell.value=str(row[int(required_columns[i])-1])
            row_num=row_num+1
        col=col+len(personalProfileColumns)

    if 'educational_details' in form:
        required_columns=form.getlist('educational_details')
        educationalDetailsColumns=[
            "Examination",
            "Name of Examination Passed",
            "Board/Institute/University",
            "Duartion of Degree in Years",
            "Status",
            "Expected Year of Passing",
            "Percentage of marks or CPI/CGPA",
            "Percent or CPI/CGPA",
            "Out of CPI/CGPA",
            "Class/Division",
            "Specialization (if any)"
        ]
        if int(required_columns[0])==0:
            required_columns=[]
            for i in range(1,len(educationalDetailsColumns)+1):
                required_columns.append(i)
        row_num=3
        maximum=0
        for query in applicantID:
            education=Education_Detail.objects.get_queryset()
            print(education)
            education.filter(applicant_id=query)
            maximum=max(maximum,len(education))
            tcol=col
            for edu in education:
                row=[
                    edu.qualification,
                    edu.examination,
                    edu.university,
                    edu.duration,
                    edu.status,
                    edu.year_of_passing,
                    edu.marks_type,
                    edu.percent,
                    edu.out_of,
                    edu.division,
                    edu.specialization
                ]
                for i in range(0,len(required_columns)):
                    cell=worksheet.cell(row=row_num,column=tcol+i)
                    cell.value=str(row[int(required_columns[i])-1])
                tcol=tcol+len(required_columns)
            row_num=row_num+1
        row_num=1
        for j in range(1,maximum+1):
            for i in range(0,len(required_columns)):
                    cell=worksheet.cell(row=row_num,column=col)
                    cell.value=str(row[int(required_columns[i])-1])
                    cell.value=str(educationalDetailsColumns[int(required_columns[i])-1])+" - "+str(j)
                    col=col+1

    if 'employment_details' in form:
        required_columns=form.getlist('employment_details')
        employmentDetailsColumns=[
            "Employment-Name of Organization ",
            "Post Held",
            "Work Type",
            "From",
            "To",
            "Period of Employment in Months",
            "Nature of Responsibilities",
            "Gross Emoluments"
        ]
        if int(required_columns[0])==0:
            required_columns=[]
            for i in range(1,len(employmentDetailsColumns)+1):
                required_columns.append(i)
        row_num=3
        maximum=0
        for query in applicantID:
            employment=Employment.objects.get_queryset()
            employment.filter(applicant_id=query)
            maximum=max(maximum,len(employment))
            tcol=col
            for emp in employment:
                row=[
                    emp.organization,
                    emp.post_held,
                    emp.work_type,
                    emp.from_date,
                    emp.to_date,
                    emp.duration,
                    emp.responsibilities,
                    emp.emoluments
                ]
                for i in range(0,len(required_columns)):
                    cell=worksheet.cell(row=row_num,column=tcol+i)
                    cell.value=str(row[int(required_columns[i])-1])
                tcol=tcol+len(required_columns)
            row_num=row_num+1
        row_num=1
        for j in range(1,maximum+1):
            for i in range(0,len(required_columns)):
                    cell=worksheet.cell(row=row_num,column=col)
                    cell.value=str(row[int(required_columns[i])-1])
                    cell.value=str(employmentDetailsColumns[int(required_columns[i])-1])+" - "+str(j)
                    col=col+1

    if 'project_details' in form:
        required_columns=form.getlist('project_details')
        projectDetailsColumns=[
            "Project Name",
            "Degree",
            "Name of University/Institute",
            "Year of Submission",
            "Name of Supervisor"
        ]
        if int(required_columns[0])==0:
            required_columns=[]
            for i in range(1,len(projectDetailsColumns)+1):
                required_columns.append(i)
        row_num=3
        maximum=0
        for query in applicantID:
            project=Project_Detail.objects.get_queryset()
            project.filter(applicant_id=query)
            maximum=max(maximum,len(project))
            tcol=col
            for pro in project:
                row=[
                    pro.title,
                    pro.degree,
                    pro.university,
                    pro.year_of_submission,
                    pro.supervisor
                ]
                for i in range(0,len(required_columns)):
                    cell=worksheet.cell(row=row_num,column=tcol+i)
                    cell.value=str(row[int(required_columns[i])-1])
                tcol=tcol+len(required_columns)
            row_num=row_num+1
        row_num=1
        for j in range(1,maximum+1):
            for i in range(0,len(required_columns)):
                    cell=worksheet.cell(row=row_num,column=col)
                    cell.value=str(row[int(required_columns[i])-1])
                    cell.value=str(projectDetailsColumns[int(required_columns[i])-1])+" - "+str(j)
                    col=col+1

    workbook.save(response)
    return response

def generate_xlsx_by_year(request, year):
    queryset = Application.objects.filter(date_applied__year=year)
    return generate_xlsx(request, queryset)
